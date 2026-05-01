from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment

# Tạo workbook và xóa sheet mặc định
wb = Workbook()
wb.remove(wb.active)

# Styles chung
header_fill = PatternFill(start_color="B4D7FF", end_color="B4D7FF", fill_type="solid")
header_font = Font(bold=True, color="000000", size=11)
title_font = Font(bold=True, size=12)
bold_font = Font(bold=True)

thin_border = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)

center_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
left_align = Alignment(horizontal='left', vertical='top', wrap_text=True)

headers = ['Column name', 'Data type', 'unsigned', 'Length/Set', 'Allow null', 'Default', 'Comment', 'Note']

# Danh sách 19 tables (đầy đủ dữ liệu từ input của bạn)
tables = [
    # 1. users
    {
        "sheet": "users",
        "title": "TABLE: users",
        "desc": "Quản lý người dùng\nLưu thông tin tài khoản người dùng của hệ thống bao gồm thông tin cá nhân, trạng thái tài khoản, vai trò và lịch sử đăng nhập",
        "data": [
            ['id', 'CHAR', '○', '10', '', 'No default', 'Primary key - ID người dùng', 'Format: 10 ký tự ngẫu nhiên'],
            ['email', 'VARCHAR', '', '255', '', 'No default', 'Email đăng nhập', 'Unique, sử dụng để xác thực'],
            ['password', 'VARCHAR', '', '255', '', 'No default', 'Mật khẩu đã mã hóa', 'Sử dụng bcrypt/hash'],
            ['full_name', 'VARCHAR', '', '255', '', 'No default', 'Họ và tên đầy đủ', ''],
            ['phone', 'VARCHAR', '', '20', '○', 'NULL', 'Số điện thoại', 'Format: +84xxxxxxxxx hoặc 0xxxxxxxxx'],
            ['avatar', 'VARCHAR', '', '255', '○', 'NULL', 'Đường dẫn ảnh đại diện', 'Lưu đường dẫn tương đối hoặc URL'],
            ['role', 'TINYINT', '', '1', '', '0', '0. member\n1. admin', 'Vai trò người dùng trong hệ thống'],
            ['account_status', 'TINYINT', '', '3', '', '0', '0. pending_approval - Chờ duyệt\n1. active - Đang hoạt động\n2. rejected - Bị từ chối\n3. suspended - Bị khóa', 'Trạng thái tài khoản người dùng'],
            ['email_verified_at', 'TIMESTAMP', '', '', '○', 'NULL', 'Thời điểm xác thực email', 'Được set khi user click link xác thực'],
            ['rejection_reason', 'VARCHAR', '', '500', '○', 'NULL', 'Lý do từ chối tài khoản', 'Chỉ có giá trị khi account_status = 2 (rejected)'],
            ['last_login_at', 'TIMESTAMP', '', '', '○', 'NULL', 'Thời điểm đăng nhập cuối', 'Cập nhật mỗi lần đăng nhập thành công'],
            ['created_at', 'TIMESTAMP', '', '', '', 'CURRENT_TIMESTAMP', 'Thời gian tạo bản ghi', ''],
            ['created_by', 'CHAR', '○', '10', '', 'No default', 'ID người tạo', 'Foreign key liên kết đến users.id'],
            ['updated_at', 'TIMESTAMP', '', '', '', 'CURRENT_TIMESTAMP', 'Thời gian cập nhật bản ghi', 'ON UPDATE CURRENT_TIMESTAMP'],
            ['updated_by', 'CHAR', '○', '10', '', 'No default', 'ID người cập nhật', 'Foreign key liên kết đến users.id'],
            ['deleted_at', 'TIMESTAMP', '', '', '○', 'NULL', 'Thời gian xóa mềm', 'Soft delete - NULL nghĩa là chưa xóa'],
            ['deleted_by', 'CHAR', '○', '10', '○', 'NULL', 'ID người xóa', 'Foreign key liên kết đến users.id'],
        ]
    },
    # 2. email_verifications
    {
        "sheet": "email_verifications",
        "title": "TABLE: email_verifications",
        "desc": "Quản lý mã xác thực email\nLưu mã OTP để xác thực email khi đăng ký hoặc đổi email",
        "data": [
            ['id', 'CHAR', '○', '10', '', 'No default', 'Primary key - ID bản ghi', 'Format: 10 ký tự ngẫu nhiên'],
            ['email', 'VARCHAR', '', '255', '', 'No default', 'Email cần xác thực', ''],
            ['verification_code', 'CHAR', '', '6', '', 'No default', 'Mã xác thực 6 số', 'Format: 123456'],
            ['expires_at', 'TIMESTAMP', '', '', '', 'No default', 'Thời điểm hết hạn', 'Thường sau 15-30 phút kể từ created_at'],
            ['is_used', 'TINYINT', '', '1', '', '0', '0. Chưa dùng\n1. Đã dùng', 'Đánh dấu mã đã được sử dụng'],
            ['used_at', 'TIMESTAMP', '', '', '○', 'NULL', 'Thời điểm sử dụng mã', 'Được set khi verify thành công'],
            ['created_at', 'TIMESTAMP', '', '', '', 'CURRENT_TIMESTAMP', 'Thời gian tạo bản ghi', ''],
            ['created_by', 'CHAR', '○', '10', '○', 'NULL', 'ID người tạo', 'Foreign key liên kết đến users.id, NULL nếu chưa có user'],
            ['updated_at', 'TIMESTAMP', '', '', '', 'CURRENT_TIMESTAMP', 'Thời gian cập nhật bản ghi', 'ON UPDATE CURRENT_TIMESTAMP'],
            ['updated_by', 'CHAR', '○', '10', '○', 'NULL', 'ID người cập nhật', 'Foreign key liên kết đến users.id'],
            ['deleted_at', 'TIMESTAMP', '', '', '○', 'NULL', 'Thời gian xóa mềm', 'Soft delete - NULL nghĩa là chưa xóa'],
            ['deleted_by', 'CHAR', '○', '10', '○', 'NULL', 'ID người xóa', 'Foreign key liên kết đến users.id'],
        ]
    },
    # 3. password_resets
    {
        "sheet": "password_resets",
        "title": "TABLE: password_resets",
        "desc": "Quản lý reset mật khẩu\nLưu token để reset mật khẩu khi người dùng quên",
        "data": [
            ['id', 'CHAR', '○', '10', '', 'No default', 'Primary key - ID bản ghi', 'Format: 10 ký tự ngẫu nhiên'],
            ['email', 'VARCHAR', '', '255', '', 'No default', 'Email cần reset mật khẩu', ''],
            ['token', 'VARCHAR', '', '255', '', 'No default', 'Token reset mật khẩu', 'Mã hóa ngẫu nhiên, gửi qua email'],
            ['expires_at', 'TIMESTAMP', '', '', '', 'No default', 'Thời điểm hết hạn', 'Thường sau 60 phút kể từ created_at'],
            ['is_used', 'TINYINT', '', '1', '', '0', '0. Chưa dùng\n1. Đã dùng', 'Đánh dấu token đã được sử dụng'],
            ['used_at', 'TIMESTAMP', '', '', '○', 'NULL', 'Thời điểm sử dụng token', 'Được set khi reset password thành công'],
            ['created_at', 'TIMESTAMP', '', '', '', 'CURRENT_TIMESTAMP', 'Thời gian tạo bản ghi', ''],
            ['created_by', 'CHAR', '○', '10', '○', 'NULL', 'ID người tạo', 'Foreign key liên kết đến users.id'],
            ['updated_at', 'TIMESTAMP', '', '', '', 'CURRENT_TIMESTAMP', 'Thời gian cập nhật bản ghi', 'ON UPDATE CURRENT_TIMESTAMP'],
            ['updated_by', 'CHAR', '○', '10', '○', 'NULL', 'ID người cập nhật', 'Foreign key liên kết đến users.id'],
            ['deleted_at', 'TIMESTAMP', '', '', '○', 'NULL', 'Thời gian xóa mềm', 'Soft delete - NULL nghĩa là chưa xóa'],
            ['deleted_by', 'CHAR', '○', '10', '○', 'NULL', 'ID người xóa', 'Foreign key liên kết đến users.id'],
        ]
    },
    # 4. user_sessions
    {
        "sheet": "user_sessions",
        "title": "TABLE: user_sessions",
        "desc": "Quản lý phiên đăng nhập\nLưu thông tin phiên đăng nhập của người dùng bao gồm access token, refresh token và thiết bị",
        "data": [
            ['id', 'CHAR', '○', '10', '', 'No default', 'Primary key - ID phiên', 'Format: 10 ký tự ngẫu nhiên'],
            ['user_id', 'CHAR', '○', '10', '', 'No default', 'ID người dùng', 'Foreign key liên kết đến users.id'],
            ['access_token', 'VARCHAR', '', '500', '', 'No default', 'JWT access token', 'Token ngắn hạn để truy cập API'],
            ['refresh_token', 'VARCHAR', '', '500', '', 'No default', 'JWT refresh token', 'Token dài hạn để làm mới access token'],
            ['device_info', 'VARCHAR', '', '255', '○', 'NULL', 'Thông tin thiết bị', 'User agent, device name'],
            ['ip_address', 'VARCHAR', '', '45', '○', 'NULL', 'Địa chỉ IP', 'Hỗ trợ IPv4 và IPv6'],
            ['expires_at', 'TIMESTAMP', '', '', '', 'No default', 'Thời điểm hết hạn', 'Thời điểm access token hết hạn'],
            ['is_active', 'TINYINT', '', '1', '', '1', '0. Không hoạt động\n1. Đang hoạt động', 'Trạng thái phiên đăng nhập'],
            ['revoked_at', 'TIMESTAMP', '', '', '○', 'NULL', 'Thời điểm thu hồi', 'Được set khi đăng xuất hoặc thu hồi token'],
            ['created_at', 'TIMESTAMP', '', '', '', 'CURRENT_TIMESTAMP', 'Thời gian tạo bản ghi', ''],
            ['created_by', 'CHAR', '○', '10', '', 'No default', 'ID người tạo', 'Foreign key liên kết đến users.id'],
            ['updated_at', 'TIMESTAMP', '', '', '', 'CURRENT_TIMESTAMP', 'Thời gian cập nhật bản ghi', 'ON UPDATE CURRENT_TIMESTAMP'],
            ['updated_by', 'CHAR', '○', '10', '○', 'NULL', 'ID người cập nhật', 'Foreign key liên kết đến users.id'],
            ['deleted_at', 'TIMESTAMP', '', '', '○', 'NULL', 'Thời gian xóa mềm', 'Soft delete - NULL nghĩa là chưa xóa'],
            ['deleted_by', 'CHAR', '○', '10', '○', 'NULL', 'ID người xóa', 'Foreign key liên kết đến users.id'],
        ]
    },
    # 5. audit_logs
    {
        "sheet": "audit_logs",
        "title": "TABLE: audit_logs",
        "desc": "Lịch sử đăng nhập và hoạt động\nGhi lại lịch sử các hành động quan trọng của người dùng trong hệ thống",
        "data": [
            ['id', 'CHAR', '○', '10', '', 'No default', 'Primary key - ID log', 'Format: 10 ký tự ngẫu nhiên'],
            ['user_id', 'CHAR', '○', '10', '○', 'NULL', 'ID người dùng', 'Foreign key liên kết đến users.id, NULL nếu action trước khi có user'],
            ['email', 'VARCHAR', '', '255', '○', 'NULL', 'Email người dùng', 'Lưu để tra cứu ngay cả khi user bị xóa'],
            ['action', 'VARCHAR', '', '100', '', 'No default', 'Hành động thực hiện', 'VD: login, logout, register, update_profile, delete_account'],
            ['ip_address', 'VARCHAR', '', '45', '○', 'NULL', 'Địa chỉ IP', 'Hỗ trợ IPv4 và IPv6'],
            ['device_info', 'VARCHAR', '', '255', '○', 'NULL', 'Thông tin thiết bị', 'User agent, device name'],
            ['details', 'VARCHAR', '', '1000', '○', 'NULL', 'Chi tiết bổ sung', 'Lưu dạng JSON hoặc text mô tả chi tiết'],
            ['created_at', 'TIMESTAMP', '', '', '', 'CURRENT_TIMESTAMP', 'Thời gian tạo bản ghi', ''],
            ['created_by', 'CHAR', '○', '10', '○', 'NULL', 'ID người tạo', 'Foreign key liên kết đến users.id'],
            ['updated_at', 'TIMESTAMP', '', '', '', 'CURRENT_TIMESTAMP', 'Thời gian cập nhật bản ghi', 'ON UPDATE CURRENT_TIMESTAMP'],
            ['updated_by', 'CHAR', '○', '10', '○', 'NULL', 'ID người cập nhật', 'Foreign key liên kết đến users.id'],
            ['deleted_at', 'TIMESTAMP', '', '', '○', 'NULL', 'Thời gian xóa mềm', 'Soft delete - NULL nghĩa là chưa xóa'],
            ['deleted_by', 'CHAR', '○', '10', '○', 'NULL', 'ID người xóa', 'Foreign key liên kết đến users.id'],
        ]
    },
    # 6. user_profiles
    {
        "sheet": "user_profiles",
        "title": "TABLE: user_profiles",
        "desc": "Thông tin cá nhân\nLưu thông tin chi tiết về sức khỏe và thể trạng của người dùng",
        "data": [
            ['id', 'CHAR', '○', '10', '', 'No default', 'Primary key - ID profile', 'Format: 10 ký tự ngẫu nhiên'],
            ['user_id', 'CHAR', '○', '10', '', 'No default', 'ID người dùng', 'Foreign key liên kết đến users.id, Unique - 1 user có 1 profile'],
            ['date_of_birth', 'DATE', '', '', '○', 'NULL', 'Ngày sinh', 'Format: YYYY-MM-DD'],
            ['gender', 'TINYINT', '', '2', '○', 'NULL', '0. male\n1. female\n2. other', 'Giới tính người dùng'],
            ['height', 'DECIMAL', '', '5,2', '○', 'NULL', 'Chiều cao (cm)', 'VD: 170.50'],
            ['current_weight', 'DECIMAL', '', '5,2', '○', 'NULL', 'Cân nặng hiện tại (kg)', 'VD: 65.50'],
            ['bmi', 'DECIMAL', '', '4,2', '○', 'NULL', 'Chỉ số BMI', 'Tính từ weight/height², VD: 22.86'],
            ['bmi_category', 'VARCHAR', '', '50', '○', 'NULL', 'Phân loại BMI', 'VD: Underweight, Normal, Overweight, Obese'],
            ['activity_level', 'TINYINT', '', '4', '', '0', '0. sedentary\n1. lightly_active\n2. moderately_active\n3. very_active\n4. extremely_active', 'Mức độ vận động hàng ngày'],
            ['created_at', 'TIMESTAMP', '', '', '', 'CURRENT_TIMESTAMP', 'Thời gian tạo bản ghi', ''],
            ['created_by', 'CHAR', '○', '10', '', 'No default', 'ID người tạo', 'Foreign key liên kết đến users.id'],
            ['updated_at', 'TIMESTAMP', '', '', '', 'CURRENT_TIMESTAMP', 'Thời gian cập nhật bản ghi', 'ON UPDATE CURRENT_TIMESTAMP'],
            ['updated_by', 'CHAR', '○', '10', '○', 'NULL', 'ID người cập nhật', 'Foreign key liên kết đến users.id'],
            ['deleted_at', 'TIMESTAMP', '', '', '○', 'NULL', 'Thời gian xóa mềm', 'Soft delete - NULL nghĩa là chưa xóa'],
            ['deleted_by', 'CHAR', '○', '10', '○', 'NULL', 'ID người xóa', 'Foreign key liên kết đến users.id'],
        ]
    },
    # 7. profile_history
    {
        "sheet": "profile_history",
        "title": "TABLE: profile_history",
        "desc": "Lịch sử thay đổi profile\nGhi lại lịch sử thay đổi các thông tin quan trọng trong profile người dùng",
        "data": [
            ['id', 'CHAR', '○', '10', '', 'No default', 'Primary key - ID history', 'Format: 10 ký tự ngẫu nhiên'],
            ['user_id', 'CHAR', '○', '10', '', 'No default', 'ID người dùng', 'Foreign key liên kết đến users.id'],
            ['field_name', 'VARCHAR', '', '100', '', 'No default', 'Tên trường thay đổi', 'VD: weight, height, activity_level'],
            ['old_value', 'VARCHAR', '', '255', '○', 'NULL', 'Giá trị cũ', 'Lưu dưới dạng string'],
            ['new_value', 'VARCHAR', '', '255', '', 'No default', 'Giá trị mới', 'Lưu dưới dạng string'],
            ['change_magnitude', 'DECIMAL', '', '10,2', '○', 'NULL', 'Độ lớn thay đổi', 'VD: Chênh lệch cân nặng: +2.5 kg'],
            ['is_significant_change', 'TINYINT', '', '1', '', '0', '0. Không đáng kể\n1. Thay đổi đáng kể', 'Đánh dấu thay đổi lớn cần chú ý'],
            ['created_at', 'TIMESTAMP', '', '', '', 'CURRENT_TIMESTAMP', 'Thời gian tạo bản ghi', ''],
            ['created_by', 'CHAR', '○', '10', '', 'No default', 'ID người tạo', 'Foreign key liên kết đến users.id'],
            ['updated_at', 'TIMESTAMP', '', '', '', 'CURRENT_TIMESTAMP', 'Thời gian cập nhật bản ghi', 'ON UPDATE CURRENT_TIMESTAMP'],
            ['updated_by', 'CHAR', '○', '10', '○', 'NULL', 'ID người cập nhật', 'Foreign key liên kết đến users.id'],
            ['deleted_at', 'TIMESTAMP', '', '', '○', 'NULL', 'Thời gian xóa mềm', 'Soft delete - NULL nghĩa là chưa xóa'],
            ['deleted_by', 'CHAR', '○', '10', '○', 'NULL', 'ID người xóa', 'Foreign key liên kết đến users.id'],
        ]
    },
    # 8. user_goals
    {
        "sheet": "user_goals",
        "title": "TABLE: user_goals",
        "desc": "Mục tiêu người dùng\nLưu mục tiêu giảm/tăng/duy trì cân nặng của người dùng",
        "data": [
            ['id', 'CHAR', '○', '10', '', 'No default', 'Primary key - ID goal', 'Format: 10 ký tự ngẫu nhiên'],
            ['user_id', 'CHAR', '○', '10', '', 'No default', 'ID người dùng', 'Foreign key liên kết đến users.id'],
            ['goal_type', 'TINYINT', '', '2', '', 'No default', '0. lose_weight\n1. gain_weight\n2. maintain_weight', 'Loại mục tiêu'],
            ['start_weight', 'DECIMAL', '', '5,2', '', 'No default', 'Cân nặng bắt đầu (kg)', 'VD: 75.50'],
            ['target_weight', 'DECIMAL', '', '5,2', '', 'No default', 'Cân nặng mục tiêu (kg)', 'VD: 65.00'],
            ['target_bmi', 'DECIMAL', '', '4,2', '○', 'NULL', 'BMI mục tiêu', 'Tính toán từ target_weight'],
            ['weekly_change_rate', 'DECIMAL', '', '4,2', '', 'No default', 'Tốc độ thay đổi/tuần (kg)', 'VD: -0.5 (giảm), +0.3 (tăng)'],
            ['estimated_weeks', 'INT', '○', '', '', 'No default', 'Số tuần dự kiến', 'Tính từ (target_weight - start_weight) / weekly_change_rate'],
            ['start_date', 'DATE', '', '', '', 'No default', 'Ngày bắt đầu', 'Format: YYYY-MM-DD'],
            ['target_date', 'DATE', '', '', '', 'No default', 'Ngày mục tiêu', 'Format: YYYY-MM-DD'],
            ['is_active', 'TINYINT', '', '1', '', '1', '0. Không hoạt động\n1. Đang hoạt động', 'Mục tiêu đang theo dõi'],
            ['is_completed', 'TINYINT', '', '1', '', '0', '0. Chưa hoàn thành\n1. Đã hoàn thành', 'Đã đạt mục tiêu'],
            ['completed_at', 'TIMESTAMP', '', '', '○', 'NULL', 'Thời điểm hoàn thành', 'Được set khi đạt target_weight'],
            ['status', 'TINYINT', '', '3', '', '0', '0. active\n1. paused\n2. completed\n3. abandoned', 'Trạng thái mục tiêu'],
            ['created_at', 'TIMESTAMP', '', '', '', 'CURRENT_TIMESTAMP', 'Thời gian tạo bản ghi', ''],
            ['created_by', 'CHAR', '○', '10', '', 'No default', 'ID người tạo', 'Foreign key liên kết đến users.id'],
            ['updated_at', 'TIMESTAMP', '', '', '', 'CURRENT_TIMESTAMP', 'Thời gian cập nhật bản ghi', 'ON UPDATE CURRENT_TIMESTAMP'],
            ['updated_by', 'CHAR', '○', '10', '○', 'NULL', 'ID người cập nhật', 'Foreign key liên kết đến users.id'],
            ['deleted_at', 'TIMESTAMP', '', '', '○', 'NULL', 'Thời gian xóa mềm', 'Soft delete - NULL nghĩa là chưa xóa'],
            ['deleted_by', 'CHAR', '○', '10', '○', 'NULL', 'ID người xóa', 'Foreign key liên kết đến users.id'],
        ]
    },
    # 9. weight_check_ins
    {
        "sheet": "weight_check_ins",
        "title": "TABLE: weight_check_ins",
        "desc": "Check-in cân nặng\nGhi lại lịch sử check-in cân nặng định kỳ của người dùng",
        "data": [
            ['id', 'CHAR', '○', '10', '', 'No default', 'Primary key - ID check-in', 'Format: 10 ký tự ngẫu nhiên'],
            ['user_id', 'CHAR', '○', '10', '', 'No default', 'ID người dùng', 'Foreign key liên kết đến users.id'],
            ['goal_id', 'CHAR', '○', '10', '○', 'NULL', 'ID mục tiêu', 'Foreign key liên kết đến user_goals.id'],
            ['weight', 'DECIMAL', '', '5,2', '', 'No default', 'Cân nặng check-in (kg)', 'VD: 68.50'],
            ['check_in_date', 'DATE', '', '', '', 'No default', 'Ngày check-in', 'Format: YYYY-MM-DD'],
            ['bmi', 'DECIMAL', '', '4,2', '○', 'NULL', 'BMI tại thời điểm check-in', 'Tính từ weight và height'],
            ['weight_change', 'DECIMAL', '', '5,2', '○', 'NULL', 'Thay đổi so với lần trước (kg)', 'VD: -1.5 (giảm), +0.8 (tăng)'],
            ['progress_percentage', 'DECIMAL', '', '5,2', '○', 'NULL', 'Tiến độ đạt mục tiêu (%)', 'Tính từ (start_weight - current_weight) / (start_weight - target_weight) * 100'],
            ['note', 'VARCHAR', '', '500', '○', 'NULL', 'Ghi chú', 'Cảm nhận, sự kiện đặc biệt'],
            ['created_at', 'TIMESTAMP', '', '', '', 'CURRENT_TIMESTAMP', 'Thời gian tạo bản ghi', ''],
            ['created_by', 'CHAR', '○', '10', '', 'No default', 'ID người tạo', 'Foreign key liên kết đến users.id'],
            ['updated_at', 'TIMESTAMP', '', '', '', 'CURRENT_TIMESTAMP', 'Thời gian cập nhật bản ghi', 'ON UPDATE CURRENT_TIMESTAMP'],
            ['updated_by', 'CHAR', '○', '10', '○', 'NULL', 'ID người cập nhật', 'Foreign key liên kết đến users.id'],
            ['deleted_at', 'TIMESTAMP', '', '', '○', 'NULL', 'Thời gian xóa mềm', 'Soft delete - NULL nghĩa là chưa xóa'],
            ['deleted_by', 'CHAR', '○', '10', '○', 'NULL', 'ID người xóa', 'Foreign key liên kết đến users.id'],
        ]
    },
    # 10. calorie_calculations
    {
        "sheet": "calorie_calculations",
        "title": "TABLE: calorie_calculations",
        "desc": "Lịch sử tính toán calorie\nLưu kết quả tính toán nhu cầu calorie và dinh dưỡng của người dùng theo thời gian",
        "data": [
            ['id', 'CHAR', '○', '10', '', 'No default', 'Primary key - ID calculation', 'Format: 10 ký tự ngẫu nhiên'],
            ['user_id', 'CHAR', '○', '10', '', 'No default', 'ID người dùng', 'Foreign key liên kết đến users.id'],
            ['goal_id', 'CHAR', '○', '10', '○', 'NULL', 'ID mục tiêu', 'Foreign key liên kết đến user_goals.id'],
            ['calculation_method', 'VARCHAR', '', '50', '', 'Mifflin-St Jeor', 'Phương pháp tính', 'VD: Mifflin-St Jeor, Harris-Benedict'],
            ['bmr', 'DECIMAL', '', '7,2', '', 'No default', 'BMR - Chuyển hóa cơ bản', 'Basal Metabolic Rate, VD: 1450.50'],
            ['tdee', 'DECIMAL', '', '7,2', '', 'No default', 'TDEE - Năng lượng tiêu hao', 'Total Daily Energy Expenditure, VD: 2175.75'],
            ['target_calories', 'DECIMAL', '', '7,2', '', 'No default', 'Calorie mục tiêu/ngày', 'TDEE +/- deficit/surplus, VD: 1675.75'],
            ['calorie_deficit_surplus', 'DECIMAL', '', '7,2', '○', 'NULL', 'Thặng dư/thâm hụt calorie', 'VD: -500 (deficit), +300 (surplus)'],
            ['protein_grams', 'DECIMAL', '', '6,2', '', 'No default', 'Protein cần thiết (gram)', 'VD: 120.50'],
            ['carbs_grams', 'DECIMAL', '', '6,2', '', 'No default', 'Carbs cần thiết (gram)', 'VD: 180.75'],
            ['fat_grams', 'DECIMAL', '', '6,2', '', 'No default', 'Fat cần thiết (gram)', 'VD: 55.25'],
            ['protein_percentage', 'DECIMAL', '', '4,2', '', 'No default', '% Protein', 'VD: 30.00'],
            ['carbs_percentage', 'DECIMAL', '', '4,2', '', 'No default', '% Carbs', 'VD: 45.00'],
            ['fat_percentage', 'DECIMAL', '', '4,2', '', 'No default', '% Fat', 'VD: 25.00'],
            ['age_at_calculation', 'INT', '○', '', '', 'No default', 'Tuổi tại thời điểm tính', 'Tính từ date_of_birth'],
            ['weight_at_calculation', 'DECIMAL', '', '5,2', '', 'No default', 'Cân nặng tại thời điểm tính (kg)', 'VD: 70.50'],
            ['height_at_calculation', 'DECIMAL', '', '5,2', '', 'No default', 'Chiều cao tại thời điểm tính (cm)', 'VD: 170.00'],
            ['activity_level', 'VARCHAR', '', '50', '', 'No default', 'Mức độ vận động', 'VD: sedentary, moderately_active'],
            ['is_warning', 'TINYINT', '', '1', '', '0', '0. Không cảnh báo\n1. Có cảnh báo', 'Đánh dấu nếu calorie quá thấp/cao'],
            ['warning_message', 'VARCHAR', '', '500', '○', 'NULL', 'Thông báo cảnh báo', 'VD: Calorie quá thấp, có thể ảnh hưởng sức khỏe'],
            ['valid_from', 'DATE', '', '', '', 'No default', 'Ngày bắt đầu hiệu lực', 'Format: YYYY-MM-DD'],
            ['valid_to', 'DATE', '', '', '○', 'NULL', 'Ngày kết thúc hiệu lực', 'NULL = vẫn còn hiệu lực'],
            ['is_active', 'TINYINT', '', '1', '', '1', '0. Không hoạt động\n1. Đang hoạt động', 'Calculation hiện đang áp dụng'],
            ['created_at', 'TIMESTAMP', '', '', '', 'CURRENT_TIMESTAMP', 'Thời gian tạo bản ghi', ''],
            ['created_by', 'CHAR', '○', '10', '', 'No default', 'ID người tạo', 'Foreign key liên kết đến users.id'],
            ['updated_at', 'TIMESTAMP', '', '', '', 'CURRENT_TIMESTAMP', 'Thời gian cập nhật bản ghi', 'ON UPDATE CURRENT_TIMESTAMP'],
            ['updated_by', 'CHAR', '○', '10', '○', 'NULL', 'ID người cập nhật', 'Foreign key liên kết đến users.id'],
            ['deleted_at', 'TIMESTAMP', '', '', '○', 'NULL', 'Thời gian xóa mềm', 'Soft delete - NULL nghĩa là chưa xóa'],
            ['deleted_by', 'CHAR', '○', '10', '○', 'NULL', 'ID người xóa', 'Foreign key liên kết đến users.id'],
        ]
    },
    # 11. food_categories
    {
        "sheet": "food_categories",
        "title": "TABLE: food_categories",
        "desc": "Danh mục thực phẩm\nPhân loại thực phẩm theo danh mục để dễ quản lý và tìm kiếm",
        "data": [
            ['id', 'CHAR', '○', '10', '', 'No default', 'Primary key - ID category', 'Format: 10 ký tự ngẫu nhiên'],
            ['name', 'VARCHAR', '', '255', '', 'No default', 'Tên danh mục', 'VD: Rau củ, Thịt, Hải sản, Trái cây'],
            ['description', 'VARCHAR', '', '500', '○', 'NULL', 'Mô tả danh mục', 'Giải thích chi tiết về danh mục'],
            ['icon', 'VARCHAR', '', '255', '○', 'NULL', 'Icon danh mục', 'Đường dẫn hoặc tên icon'],
            ['sort_order', 'INT', '○', '', '', '0', 'Thứ tự sắp xếp', 'Số nhỏ hơn hiển thị trước'],
            ['status', 'TINYINT', '', '1', '', '1', '0. Không hoạt động\n1. Đang hoạt động', 'Trạng thái danh mục'],
            ['created_at', 'TIMESTAMP', '', '', '', 'CURRENT_TIMESTAMP', 'Thời gian tạo bản ghi', ''],
            ['created_by', 'CHAR', '○', '10', '', 'No default', 'ID người tạo', 'Foreign key liên kết đến users.id'],
            ['updated_at', 'TIMESTAMP', '', '', '', 'CURRENT_TIMESTAMP', 'Thời gian cập nhật bản ghi', 'ON UPDATE CURRENT_TIMESTAMP'],
            ['updated_by', 'CHAR', '○', '10', '○', 'NULL', 'ID người cập nhật', 'Foreign key liên kết đến users.id'],
            ['deleted_at', 'TIMESTAMP', '', '', '○', 'NULL', 'Thời gian xóa mềm', 'Soft delete - NULL nghĩa là chưa xóa'],
            ['deleted_by', 'CHAR', '○', '10', '○', 'NULL', 'ID người xóa', 'Foreign key liên kết đến users.id'],
        ]
    },
    # 12. foods
    {
        "sheet": "foods",
        "title": "TABLE: foods",
        "desc": "Danh sách thực phẩm/món ăn\nLưu thông tin chi tiết về thực phẩm, món ăn và giá trị dinh dưỡng",
        "data": [
            ['id', 'CHAR', '○', '10', '', 'No default', 'Primary key - ID food', 'Format: 10 ký tự ngẫu nhiên'],
            ['category_id', 'CHAR', '○', '10', '○', 'NULL', 'ID danh mục', 'Foreign key liên kết đến food_categories.id'],
            ['name', 'VARCHAR', '', '255', '', 'No default', 'Tên thực phẩm/món ăn', 'VD: Cơm trắng, Gà luộc'],
            ['description', 'VARCHAR', '', '500', '○', 'NULL', 'Mô tả chi tiết', 'Thành phần, cách chế biến'],
            ['image', 'VARCHAR', '', '255', '○', 'NULL', 'Hình ảnh món ăn', 'Đường dẫn tương đối hoặc URL'],
            ['serving_size', 'DECIMAL', '', '7,2', '', 'No default', 'Khối lượng 1 phần ăn (gram)', 'VD: 100.00 (100g)'],
            ['serving_unit', 'VARCHAR', '', '50', '', 'gram', 'Đơn vị tính', 'VD: gram, ml, chén, bát'],
            ['calories', 'DECIMAL', '', '7,2', '', 'No default', 'Năng lượng (kcal)', 'VD: 130.00 kcal/100g'],
            ['protein', 'DECIMAL', '', '6,2', '', 'No default', 'Protein (gram)', 'VD: 25.50 g/100g'],
            ['carbs', 'DECIMAL', '', '6,2', '', 'No default', 'Carbohydrate (gram)', 'VD: 45.30 g/100g'],
            ['fat', 'DECIMAL', '', '6,2', '', 'No default', 'Chất béo (gram)', 'VD: 8.20 g/100g'],
            ['fiber', 'DECIMAL', '', '6,2', '○', 'NULL', 'Chất xơ (gram)', 'VD: 2.50 g/100g'],
            ['is_vegetarian', 'TINYINT', '', '1', '', '0', '0. Không\n1. Chay', 'Món ăn chay'],
            ['is_vegan', 'TINYINT', '', '1', '', '0', '0. Không\n1. Thuần chay', 'Món ăn thuần chay (không trứng, sữa)'],
            ['meal_type', 'VARCHAR', '', '100', '○', 'NULL', 'Loại bữa ăn phù hợp', 'breakfast,lunch,dinner,snack - lưu dạng text phân cách bằng dấu phẩy'],
            ['preparation_time', 'INT', '○', '', '○', 'NULL', 'Thời gian chuẩn bị (phút)', 'VD: 30 (phút)'],
            ['popularity_score', 'INT', '○', '', '', '0', 'Điểm phổ biến', 'Tăng mỗi khi được chọn, dùng để gợi ý'],
            ['status', 'TINYINT', '', '1', '', '1', '0. Không hoạt động\n1. Đang hoạt động', 'Trạng thái món ăn'],
            ['created_at', 'TIMESTAMP', '', '', '', 'CURRENT_TIMESTAMP', 'Thời gian tạo bản ghi', ''],
            ['created_by', 'CHAR', '○', '10', '', 'No default', 'ID người tạo', 'Foreign key liên kết đến users.id'],
            ['updated_at', 'TIMESTAMP', '', '', '', 'CURRENT_TIMESTAMP', 'Thời gian cập nhật bản ghi', 'ON UPDATE CURRENT_TIMESTAMP'],
            ['updated_by', 'CHAR', '○', '10', '○', 'NULL', 'ID người cập nhật', 'Foreign key liên kết đến users.id'],
            ['deleted_at', 'TIMESTAMP', '', '', '○', 'NULL', 'Thời gian xóa mềm', 'Soft delete - NULL nghĩa là chưa xóa'],
            ['deleted_by', 'CHAR', '○', '10', '○', 'NULL', 'ID người xóa', 'Foreign key liên kết đến users.id'],
        ]
    },
    # 13. food_ingredients
    {
        "sheet": "food_ingredients",
        "title": "TABLE: food_ingredients",
        "desc": "Thành phần món ăn\nLưu chi tiết nguyên liệu cấu thành món ăn",
        "data": [
            ['id', 'CHAR', '○', '10', '', 'No default', 'Primary key - ID ingredient', 'Format: 10 ký tự ngẫu nhiên'],
            ['food_id', 'CHAR', '○', '10', '', 'No default', 'ID món ăn', 'Foreign key liên kết đến foods.id'],
            ['ingredient_name', 'VARCHAR', '', '255', '', 'No default', 'Tên nguyên liệu', 'VD: Thịt gà, Rau cải, Dầu ăn'],
            ['quantity', 'DECIMAL', '', '7,2', '', 'No default', 'Số lượng', 'VD: 200.00'],
            ['unit', 'VARCHAR', '', '50', '', 'No default', 'Đơn vị tính', 'VD: gram, ml, muỗng canh'],
            ['created_at', 'TIMESTAMP', '', '', '', 'CURRENT_TIMESTAMP', 'Thời gian tạo bản ghi', ''],
            ['created_by', 'CHAR', '○', '10', '', 'No default', 'ID người tạo', 'Foreign key liên kết đến users.id'],
            ['updated_at', 'TIMESTAMP', '', '', '', 'CURRENT_TIMESTAMP', 'Thời gian cập nhật bản ghi', 'ON UPDATE CURRENT_TIMESTAMP'],
            ['updated_by', 'CHAR', '○', '10', '○', 'NULL', 'ID người cập nhật', 'Foreign key liên kết đến users.id'],
            ['deleted_at', 'TIMESTAMP', '', '', '○', 'NULL', 'Thời gian xóa mềm', 'Soft delete - NULL nghĩa là chưa xóa'],
            ['deleted_by', 'CHAR', '○', '10', '○', 'NULL', 'ID người xóa', 'Foreign key liên kết đến users.id'],
        ]
    },
    # 14. meal_plans
    {
        "sheet": "meal_plans",
        "title": "TABLE: meal_plans",
        "desc": "Kế hoạch bữa ăn\nLưu kế hoạch bữa ăn theo tuần/tháng của người dùng",
        "data": [
            ['id', 'CHAR', '○', '10', '', 'No default', 'Primary key - ID meal plan', 'Format: 10 ký tự ngẫu nhiên'],
            ['user_id', 'CHAR', '○', '10', '', 'No default', 'ID người dùng', 'Foreign key liên kết đến users.id'],
            ['goal_id', 'CHAR', '○', '10', '○', 'NULL', 'ID mục tiêu', 'Foreign key liên kết đến user_goals.id'],
            ['calorie_calculation_id', 'CHAR', '○', '10', '○', 'NULL', 'ID calculation', 'Foreign key liên kết đến calorie_calculations.id'],
            ['plan_name', 'VARCHAR', '', '255', '', 'No default', 'Tên kế hoạch', 'VD: Kế hoạch tuần 1, Giảm cân tháng 5'],
            ['start_date', 'DATE', '', '', '', 'No default', 'Ngày bắt đầu', 'Format: YYYY-MM-DD'],
            ['end_date', 'DATE', '', '', '', 'No default', 'Ngày kết thúc', 'Format: YYYY-MM-DD'],
            ['total_days', 'INT', '○', '', '', 'No default', 'Tổng số ngày', 'Tính từ end_date - start_date + 1'],
            ['target_calories_per_day', 'DECIMAL', '', '7,2', '', 'No default', 'Calorie mục tiêu/ngày', 'VD: 1800.00'],
            ['is_auto_generated', 'TINYINT', '', '1', '', '1', '0. Tạo thủ công\n1. Tự động tạo', 'Kế hoạch do hệ thống gợi ý'],
            ['is_active', 'TINYINT', '', '1', '', '1', '0. Không hoạt động\n1. Đang hoạt động', 'Kế hoạch đang theo dõi'],
            ['status', 'TINYINT', '', '3', '', '1', '0. draft\n1. active\n2. completed\n3. cancelled', 'Trạng thái kế hoạch'],
            ['created_at', 'TIMESTAMP', '', '', '', 'CURRENT_TIMESTAMP', 'Thời gian tạo bản ghi', ''],
            ['created_by', 'CHAR', '○', '10', '', 'No default', 'ID người tạo', 'Foreign key liên kết đến users.id'],
            ['updated_at', 'TIMESTAMP', '', '', '', 'CURRENT_TIMESTAMP', 'Thời gian cập nhật bản ghi', 'ON UPDATE CURRENT_TIMESTAMP'],
            ['updated_by', 'CHAR', '○', '10', '○', 'NULL', 'ID người cập nhật', 'Foreign key liên kết đến users.id'],
            ['deleted_at', 'TIMESTAMP', '', '', '○', 'NULL', 'Thời gian xóa mềm', 'Soft delete - NULL nghĩa là chưa xóa'],
            ['deleted_by', 'CHAR', '○', '10', '○', 'NULL', 'ID người xóa', 'Foreign key liên kết đến users.id'],
        ]
    },
    # 15. meal_plan_details
    {
        "sheet": "meal_plan_details",
        "title": "TABLE: meal_plan_details",
        "desc": "Chi tiết bữa ăn\nLưu chi tiết món ăn trong từng bữa của kế hoạch",
        "data": [
            ['id', 'CHAR', '○', '10', '', 'No default', 'Primary key - ID detail', 'Format: 10 ký tự ngẫu nhiên'],
            ['meal_plan_id', 'CHAR', '○', '10', '', 'No default', 'ID kế hoạch', 'Foreign key liên kết đến meal_plans.id'],
            ['day_number', 'INT', '○', '', '', 'No default', 'Ngày thứ mấy trong kế hoạch', 'VD: 1, 2, 3... (1-7 cho kế hoạch tuần)'],
            ['meal_type', 'TINYINT', '', '3', '', 'No default', '0. breakfast\n1. lunch\n2. dinner\n3. snack', 'Loại bữa ăn'],
            ['food_id', 'CHAR', '○', '10', '', 'No default', 'ID món ăn', 'Foreign key liên kết đến foods.id'],
            ['serving_size', 'DECIMAL', '', '7,2', '', 'No default', 'Khối lượng 1 phần (gram)', 'VD: 100.00'],
            ['servings', 'DECIMAL', '', '5,2', '', '1.00', 'Số phần ăn', 'VD: 1.5 (1.5 phần)'],
            ['total_calories', 'DECIMAL', '', '7,2', '', 'No default', 'Tổng calories', '= calories × servings'],
            ['total_protein', 'DECIMAL', '', '6,2', '', 'No default', 'Tổng protein (gram)', '= protein × servings'],
            ['total_carbs', 'DECIMAL', '', '6,2', '', 'No default', 'Tổng carbs (gram)', '= carbs × servings'],
            ['total_fat', 'DECIMAL', '', '6,2', '', 'No default', 'Tổng fat (gram)', '= fat × servings'],
            ['created_at', 'TIMESTAMP', '', '', '', 'CURRENT_TIMESTAMP', 'Thời gian tạo bản ghi', ''],
            ['created_by', 'CHAR', '○', '10', '', 'No default', 'ID người tạo', 'Foreign key liên kết đến users.id'],
            ['updated_at', 'TIMESTAMP', '', '', '', 'CURRENT_TIMESTAMP', 'Thời gian cập nhật bản ghi', 'ON UPDATE CURRENT_TIMESTAMP'],
            ['updated_by', 'CHAR', '○', '10', '○', 'NULL', 'ID người cập nhật', 'Foreign key liên kết đến users.id'],
            ['deleted_at', 'TIMESTAMP', '', '', '○', 'NULL', 'Thời gian xóa mềm', 'Soft delete - NULL nghĩa là chưa xóa'],
            ['deleted_by', 'CHAR', '○', '10', '○', 'NULL', 'ID người xóa', 'Foreign key liên kết đến users.id'],
        ]
    },
    # 16. nutrition_logs
    {
        "sheet": "nutrition_logs",
        "title": "TABLE: nutrition_logs",
        "desc": "Nhật ký dinh dưỡng\nGhi lại thực tế các món ăn người dùng đã tiêu thụ hàng ngày",
        "data": [
            ['id', 'CHAR', '○', '10', '', 'No default', 'Primary key - ID log', 'Format: 10 ký tự ngẫu nhiên'],
            ['user_id', 'CHAR', '○', '10', '', 'No default', 'ID người dùng', 'Foreign key liên kết đến users.id'],
            ['log_date', 'DATE', '', '', '', 'No default', 'Ngày ghi nhật ký', 'Format: YYYY-MM-DD'],
            ['meal_type', 'TINYINT', '', '3', '', 'No default', '0. breakfast\n1. lunch\n2. dinner\n3. snack', 'Loại bữa ăn'],
            ['food_id', 'CHAR', '○', '10', '○', 'NULL', 'ID món ăn', 'Foreign key liên kết đến foods.id, NULL nếu là món tự nhập'],
            ['custom_food_name', 'VARCHAR', '', '255', '○', 'NULL', 'Tên món tự nhập', 'Dùng khi food_id = NULL'],
            ['serving_size', 'DECIMAL', '', '7,2', '', 'No default', 'Khối lượng 1 phần (gram)', 'VD: 100.00'],
            ['servings', 'DECIMAL', '', '5,2', '', '1.00', 'Số phần ăn', 'VD: 1.5 (1.5 phần)'],
            ['calories', 'DECIMAL', '', '7,2', '', 'No default', 'Tổng calories', '= calories × servings'],
            ['protein', 'DECIMAL', '', '6,2', '', 'No default', 'Tổng protein (gram)', '= protein × servings'],
            ['carbs', 'DECIMAL', '', '6,2', '', 'No default', 'Tổng carbs (gram)', '= carbs × servings'],
            ['fat', 'DECIMAL', '', '6,2', '', 'No default', 'Tổng fat (gram)', '= fat × servings'],
            ['is_flagged_unrealistic', 'TINYINT', '', '1', '', '0', '0. Bình thường\n1. Không hợp lý', 'Đánh dấu giá trị bất thường'],
            ['note', 'VARCHAR', '', '500', '○', 'NULL', 'Ghi chú', 'Cảm nhận, địa điểm ăn'],
            ['created_at', 'TIMESTAMP', '', '', '', 'CURRENT_TIMESTAMP', 'Thời gian tạo bản ghi', ''],
            ['created_by', 'CHAR', '○', '10', '', 'No default', 'ID người tạo', 'Foreign key liên kết đến users.id'],
            ['updated_at', 'TIMESTAMP', '', '', '', 'CURRENT_TIMESTAMP', 'Thời gian cập nhật bản ghi', 'ON UPDATE CURRENT_TIMESTAMP'],
            ['updated_by', 'CHAR', '○', '10', '○', 'NULL', 'ID người cập nhật', 'Foreign key liên kết đến users.id'],
            ['deleted_at', 'TIMESTAMP', '', '', '○', 'NULL', 'Thời gian xóa mềm', 'Soft delete - NULL nghĩa là chưa xóa'],
            ['deleted_by', 'CHAR', '○', '10', '○', 'NULL', 'ID người xóa', 'Foreign key liên kết đến users.id'],
        ]
    },
    # 17. daily_nutrition_summary
    {
        "sheet": "daily_nutrition_summary",
        "title": "TABLE: daily_nutrition_summary",
        "desc": "Tổng hợp dinh dưỡng theo ngày\nTổng hợp tất cả nutrition logs của người dùng theo ngày để theo dõi tiến độ",
        "data": [
            ['id', 'CHAR', '○', '10', '', 'No default', 'Primary key - ID summary', 'Format: 10 ký tự ngẫu nhiên'],
            ['user_id', 'CHAR', '○', '10', '', 'No default', 'ID người dùng', 'Foreign key liên kết đến users.id'],
            ['summary_date', 'DATE', '', '', '', 'No default', 'Ngày tổng hợp', 'Format: YYYY-MM-DD, Unique per user'],
            ['target_calories', 'DECIMAL', '', '7,2', '', 'No default', 'Calorie mục tiêu', 'Lấy từ calorie_calculations.target_calories'],
            ['consumed_calories', 'DECIMAL', '', '7,2', '', '0.00', 'Calorie đã tiêu thụ', 'Tổng từ nutrition_logs'],
            ['remaining_calories', 'DECIMAL', '', '7,2', '', 'No default', 'Calorie còn lại', '= target_calories - consumed_calories'],
            ['target_protein', 'DECIMAL', '', '6,2', '', 'No default', 'Protein mục tiêu (gram)', 'Lấy từ calorie_calculations'],
            ['consumed_protein', 'DECIMAL', '', '6,2', '', '0.00', 'Protein đã tiêu thụ (gram)', 'Tổng từ nutrition_logs'],
            ['target_carbs', 'DECIMAL', '', '6,2', '', 'No default', 'Carbs mục tiêu (gram)', 'Lấy từ calorie_calculations'],
            ['consumed_carbs', 'DECIMAL', '', '6,2', '', '0.00', 'Carbs đã tiêu thụ (gram)', 'Tổng từ nutrition_logs'],
            ['target_fat', 'DECIMAL', '', '6,2', '', 'No default', 'Fat mục tiêu (gram)', 'Lấy từ calorie_calculations'],
            ['consumed_fat', 'DECIMAL', '', '6,2', '', '0.00', 'Fat đã tiêu thụ (gram)', 'Tổng từ nutrition_logs'],
            ['is_goal_met', 'TINYINT', '', '1', '', '0', '0. Chưa đạt\n1. Đã đạt', 'Đã đạt mục tiêu calorie hôm nay'],
            ['variance_percentage', 'DECIMAL', '', '6,2', '○', 'NULL', '% Chênh lệch so với mục tiêu', '= (consumed - target) / target × 100'],
            ['created_at', 'TIMESTAMP', '', '', '', 'CURRENT_TIMESTAMP', 'Thời gian tạo bản ghi', ''],
            ['created_by', 'CHAR', '○', '10', '', 'No default', 'ID người tạo', 'Foreign key liên kết đến users.id'],
            ['updated_at', 'TIMESTAMP', '', '', '', 'CURRENT_TIMESTAMP', 'Thời gian cập nhật bản ghi', 'ON UPDATE CURRENT_TIMESTAMP'],
            ['updated_by', 'CHAR', '○', '10', '○', 'NULL', 'ID người cập nhật', 'Foreign key liên kết đến users.id'],
            ['deleted_at', 'TIMESTAMP', '', '', '○', 'NULL', 'Thời gian xóa mềm', 'Soft delete - NULL nghĩa là chưa xóa'],
            ['deleted_by', 'CHAR', '○', '10', '○', 'NULL', 'ID người xóa', 'Foreign key liên kết đến users.id'],
        ]
    },
    # 18. user_food_history
    {
        "sheet": "user_food_history",
        "title": "TABLE: user_food_history",
        "desc": "Lịch sử thực phẩm thường dùng\nLưu lịch sử món ăn người dùng hay sử dụng để gợi ý",
        "data": [
            ['id', 'CHAR', '○', '10', '', 'No default', 'Primary key - ID history', 'Format: 10 ký tự ngẫu nhiên'],
            ['user_id', 'CHAR', '○', '10', '', 'No default', 'ID người dùng', 'Foreign key liên kết đến users.id'],
            ['food_id', 'CHAR', '○', '10', '', 'No default', 'ID món ăn', 'Foreign key liên kết đến foods.id'],
            ['frequency_count', 'INT', '○', '', '', '1', 'Số lần sử dụng', 'Tăng mỗi khi user chọn món này'],
            ['last_consumed_at', 'TIMESTAMP', '', '', '', 'No default', 'Lần tiêu thụ gần nhất', 'Cập nhật mỗi khi user log món này'],
            ['created_at', 'TIMESTAMP', '', '', '', 'CURRENT_TIMESTAMP', 'Thời gian tạo bản ghi', ''],
            ['created_by', 'CHAR', '○', '10', '', 'No default', 'ID người tạo', 'Foreign key liên kết đến users.id'],
            ['updated_at', 'TIMESTAMP', '', '', '', 'CURRENT_TIMESTAMP', 'Thời gian cập nhật bản ghi', 'ON UPDATE CURRENT_TIMESTAMP'],
            ['updated_by', 'CHAR', '○', '10', '○', 'NULL', 'ID người cập nhật', 'Foreign key liên kết đến users.id'],
            ['deleted_at', 'TIMESTAMP', '', '', '○', 'NULL', 'Thời gian xóa mềm', 'Soft delete - NULL nghĩa là chưa xóa'],
            ['deleted_by', 'CHAR', '○', '10', '○', 'NULL', 'ID người xóa', 'Foreign key liên kết đến users.id'],
        ]
    },
    # 19. notifications
    {
        "sheet": "notifications",
        "title": "TABLE: notifications",
        "desc": "Thông báo\nQuản lý thông báo gửi đến người dùng",
        "data": [
            ['id', 'CHAR', '○', '10', '', 'No default', 'Primary key - ID notification', 'Format: 10 ký tự ngẫu nhiên'],
            ['user_id', 'CHAR', '○', '10', '', 'No default', 'ID người dùng', 'Foreign key liên kết đến users.id'],
            ['title', 'VARCHAR', '', '255', '', 'No default', 'Tiêu đề thông báo', 'VD: Nhắc nhở check-in cân nặng'],
            ['message', 'VARCHAR', '', '500', '', 'No default', 'Nội dung thông báo', 'Chi tiết thông báo'],
            ['type', 'TINYINT', '', '3', '', 'No default', '0. reminder\n1. alert\n2. achievement\n3. system', 'Loại thông báo'],
            ['is_read', 'TINYINT', '', '1', '', '0', '0. Chưa đọc\n1. Đã đọc', 'Trạng thái đọc'],
            ['read_at', 'TIMESTAMP', '', '', '○', 'NULL', 'Thời điểm đọc', 'Được set khi user mở thông báo'],
            ['created_at', 'TIMESTAMP', '', '', '', 'CURRENT_TIMESTAMP', 'Thời gian tạo bản ghi', ''],
            ['created_by', 'CHAR', '○', '10', '', 'No default', 'ID người tạo', 'Foreign key liên kết đến users.id'],
            ['updated_at', 'TIMESTAMP', '', '', '', 'CURRENT_TIMESTAMP', 'Thời gian cập nhật bản ghi', 'ON UPDATE CURRENT_TIMESTAMP'],
            ['updated_by', 'CHAR', '○', '10', '○', 'NULL', 'ID người cập nhật', 'Foreign key liên kết đến users.id'],
            ['deleted_at', 'TIMESTAMP', '', '', '○', 'NULL', 'Thời gian xóa mềm', 'Soft delete - NULL nghĩa là chưa xóa'],
            ['deleted_by', 'CHAR', '○', '10', '○', 'NULL', 'ID người xóa', 'Foreign key liên kết đến users.id'],
        ]
    },
]

# Hàm tạo sheet
def create_sheet(table_info):
    ws = wb.create_sheet(title=table_info["sheet"])

    # Row 1-2: Title
    ws['B2'] = table_info["title"]
    ws['B2'].font = title_font

    # Row 3: Description (merge B3:I3)
    ws['B3'] = table_info["desc"]
    ws.merge_cells('B3:I3')
    ws['B3'].alignment = left_align
    ws.row_dimensions[3].height = 35

    # Row 5: Table name
    ws['B5'] = table_info["sheet"]
    ws['B5'].font = bold_font
    ws['B5'].alignment = left_align

    # Row 6: Headers
    for col_idx, header in enumerate(headers, start=2):
        cell = ws.cell(row=6, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = thin_border
        cell.alignment = center_align

    # Data rows từ row 7
    for row_idx, row_data in enumerate(table_info["data"], start=7):
        for col_idx, value in enumerate(row_data, start=2):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.border = thin_border
            if col_idx in [8, 9]:  # Comment và Note: left align
                cell.alignment = left_align
            else:
                cell.alignment = center_align

        # Tăng chiều cao dòng nếu có nội dung dài (có \n)
        if any('\n' in str(value) for value in row_data):
            ws.row_dimensions[row_idx].height = 65

    # Độ rộng cột cố định
    column_widths = {'B': 20, 'C': 15, 'D': 10, 'E': 15, 'F': 12, 'G': 18, 'H': 25, 'I': 35}
    for col_letter, width in column_widths.items():
        ws.column_dimensions[col_letter].width = width

# Tạo tất cả 19 sheets
for table in tables:
    create_sheet(table)

# Lưu file
output_file = 'database_design_19_tables.xlsx'
wb.save(output_file)
print(f"✅ Đã tạo file: {output_file}")
print("🎉 Hoàn thành! Mở file Excel để xem 19 sheet với format đẹp.")

if __name__ == "__main__":
    pass  # Code đã chạy ở trên
